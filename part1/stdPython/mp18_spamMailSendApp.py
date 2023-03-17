# 대량 메일 전송

from openpyxl import load_workbook # 엑셀을 파이썬에서 핸들링하기 위한 모듈 # pip install openpyxl
import smtplib          # 메일 전송 프로토콜
from email.mime.text import MIMEText        # MIME: Multipurpose Internet Mail Extension
from email.mime.multipart import MIMEMultipart

wbook = load_workbook('./stdPython/SpamMailList.xlsx', data_only=True)
wsheet = wbook.active         # sheet1 선택

for i in range(1, wsheet.max_row+1):
    recv_email = wsheet.cell(i, 1).value
    print(recv_email)

    try:
        # 실제 메일 전송 로직
        send_email = '2000goal@naver.com'
        send_pass = 'limshepwd'        # 임시 비밀번호
        smtp_name = 'smtp.naver.com'
        smtp_port = 587     # 포트번호

        text = '''이 편지는 영국에서 최초로 시작되어 일년에 한바퀴를 돌면서 받는 사람에게 행운을 주었고 지금은 당신에게로 옮겨진 이 편지는 4일 안에 당신 곁을 떠나야 합니다. 이 편지를 포함해서 7통을 행운이 필요한 사람에게 보내 주셔야 합니다. 복사를 해도 좋습니다. 혹 미신이라 하실지 모르지만 사실입니다. 영국에서 HGXWCH이라는 사람은 1930년에 이 편지를 받았습니다. 그는 비서에게 복사해서 보내라고 했습니다. 며칠 뒤에 복권이 당첨되어 20억을 받았습니다. 어떤 이는 이 편지를 받았으나 96시간 이내 자신의 손에서 떠나야 한다는 사실을 잊었습니다. 그는 곧 사직되었습니다. 나중에야 이 사실을 알고 7통의 편지를 보냈는데 다시 좋은 직장을 얻었습니다. 미국의 케네디 대통령은 이 편지를 받았지만 그냥 버렸습니다. 결국 9일 후 그는 암살당했습니다. 기억해 주세요. 이 편지를 보내면 7년의 행운이 있을 것이고 그렇지 않으면 3년의 불행이 있을 것입니다. 그리고 이 편지를 버리거나 낙서를 해서는 절대로 안됩니다. 7통입니다. 이 편지를 받은 사람은 행운이 깃들것입니다. 힘들겠지만 좋은게 좋다고 생각하세요. 7년의 행운을 빌면서...'''

        msg = MIMEMultipart()
        msg['Subject'] = '엑셀에서 보내는 메일입니다'
        msg['From'] = send_email    # 보내는 메일
        msg['To'] = recv_email      # 받는 메일
        msg.attach(MIMEText(text))

        mail = smtplib.SMTP(smtp_name, smtp_port)   # SMTP 객체생성
        mail.starttls()             # 전송계층 보안 시작
        mail.login(send_email, send_pass)
        mail.sendmail(send_email, recv_email, msg=msg.as_string())
        mail.quit()
        print('전송성공!')
    except Exception as e:
        print(f'수신메일 {recv_email}')
        print(f'전송에러: {e}')

